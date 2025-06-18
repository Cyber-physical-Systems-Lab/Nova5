import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

emotion_data_path = "emotion_data.xlsx"  # Path to the Excel file with emotion data

# Load emotional data from Excel sheet
emotion_df = pd.read_excel(emotion_data_path)
folder_path = os.getcwd()  # Get the current working directory
# Find the experiment data file in the current directory    
exp_data_path = next((f for f in os.listdir(folder_path) if f.startswith("P") and f.endswith(".xlsx")), None)
if exp_data_path is None:
    raise FileNotFoundError("No .xlsx file starting with 'P' found in the current folder.")
exp_data_path = os.path.join(folder_path, exp_data_path)

# Load experiment data
exp_df = pd.read_excel(exp_data_path)

# Define block start and end indices
block_ranges = {
    1: (15, 31),
    2: (34, 50),
    3: (53, 69)
}

# Define frame rate (fps) of the video recording
fps = 15  # adjust this value according to your video setup

# Define emotion columns
emotion_columns = ['Angry', 'Disgust', 'Fear', 'Happy', 'Neutral', 'Sad', 'Surprise']

# Function to process a block
def process_block(block_number, block_start, block_end):
    # Extract operation times, cube colors, and correctness values
    operation_times = exp_df.iloc[block_start:block_end, 1].astype(float).tolist()
    cube_color = exp_df.iloc[block_start:block_end, 2].astype(str).tolist()  # Green or Red
    trial_correctness = exp_df.iloc[block_start:block_end, 3].astype(str).tolist()  # direct or indirect
    true_false_values = exp_df.iloc[block_start:block_end, 4].astype(str).tolist()  # True or False

    # Compute cumulative frame ranges for each trial
    total_frames = [int(t * fps) for t in operation_times]
    start_frames = np.cumsum([0] + total_frames[:-1])
    end_frames = np.cumsum(total_frames)

    # Slice emotion data per trial and summarize dominant emotion distribution
    trial_emotions = []
    for i, (start, end) in enumerate(zip(start_frames, end_frames)):
        trial_slice = emotion_df.iloc[start:end]
        dominant_series = trial_slice[emotion_columns].apply(
            lambda row: row.idxmax() if row.notna().any() else None, axis=1
        )

        # Filter out rows where 'Dominant' is None
        valid_rows = dominant_series.dropna()

        # Calculate percentages
        dominant_counts = valid_rows.value_counts(normalize=True) * 100  # Percentage of valid rows
        summary = {emotion: dominant_counts.get(emotion, 0.0) for emotion in emotion_columns}

        summary['Trial'] = i + 1
        trial_emotions.append(summary)

    # Create DataFrame for analysis
    trial_emotion_df = pd.DataFrame(trial_emotions)

    # Move Trial column to the first position
    trial_emotion_df = trial_emotion_df[['Trial'] + emotion_columns]

    # Add cube color and correctness information
    trial_emotion_df['Cube Color'] = cube_color
    trial_emotion_df['Correctness'] = trial_correctness
    trial_emotion_df['True/False'] = true_false_values

    # Perform Early/Late analysis
    half_point = len(trial_emotion_df) // 2
    early_trials = trial_emotion_df.iloc[:half_point]
    late_trials = trial_emotion_df.iloc[half_point:]

    # Perform Green/Red trial analysis
    green_trials = trial_emotion_df[trial_emotion_df['Cube Color'] == 'Green']
    red_trials = trial_emotion_df[trial_emotion_df['Cube Color'] == 'Red']

    # Perform True/False analysis
    true_trials = trial_emotion_df[trial_emotion_df['True/False'] == 'True']
    false_trials = trial_emotion_df[trial_emotion_df['True/False'] == 'False']


    # Perform Green/Red Early/Late analysis
    green_early_trials = early_trials[early_trials['Cube Color'] == 'Green']
    red_early_trials = early_trials[early_trials['Cube Color'] == 'Red']
    green_late_trials = late_trials[late_trials['Cube Color'] == 'Green']
    red_late_trials = late_trials[late_trials['Cube Color'] == 'Red']
    early_emotion_distribution = early_trials[emotion_columns].mean()
    late_emotion_distribution = late_trials[emotion_columns].mean()
    green_emotion_distribution = green_trials[emotion_columns].mean()
    red_emotion_distribution = red_trials[emotion_columns].mean()
    green_early_distribution = green_early_trials[emotion_columns].mean()
    red_early_distribution = red_early_trials[emotion_columns].mean()
    green_late_distribution = green_late_trials[emotion_columns].mean()
    red_late_distribution = red_late_trials[emotion_columns].mean()

    # Add Green/Red analysis as the last row
    green_row = pd.Series({'Trial': 'Green Trials', **green_emotion_distribution})
    red_row = pd.Series({'Trial': 'Red Trials', **red_emotion_distribution})
    early_row = pd.Series({'Trial': 'Early Trials', **early_emotion_distribution})
    late_row = pd.Series({'Trial': 'Late Trials', **late_emotion_distribution})
    green_early_row = pd.Series({'Trial': 'Green Early Trials', **green_early_distribution})
    red_early_row = pd.Series({'Trial': 'Red Early Trials', **red_early_distribution})
    green_late_row = pd.Series({'Trial': 'Green Late Trials', **green_late_distribution})
    red_late_row = pd.Series({'Trial': 'Red Late Trials', **red_late_distribution})

    trial_emotion_df = pd.concat([trial_emotion_df, pd.DataFrame([green_row, red_row, early_row, late_row,green_early_row, red_early_row, green_late_row, red_late_row])], ignore_index=True)

    # Perform Green/Red Direct/Indirect Early analysis for Block 2
    if block_number == 2:
        direct_trials = trial_emotion_df[trial_emotion_df['Correctness'] == 'direct']
        indirect_trials = trial_emotion_df[trial_emotion_df['Correctness'] == 'indirect']
        green_direct_early_trials = green_early_trials[green_early_trials['Correctness'] == 'direct']
        red_direct_early_trials = red_early_trials[red_early_trials['Correctness'] == 'direct']
        green_indirect_early_trials = green_early_trials[green_early_trials['Correctness'] == 'indirect']
        red_indirect_early_trials = red_early_trials[red_early_trials['Correctness'] == 'indirect']
        green_direct_late_trials = green_late_trials[green_late_trials['Correctness'] == 'direct']
        red_direct_late_trials = red_late_trials[red_late_trials['Correctness'] == 'direct']
        green_indirect_late_trials = green_late_trials[green_late_trials['Correctness'] == 'indirect']
        red_indirect_late_trials = red_late_trials[red_late_trials['Correctness'] == 'indirect']

        # Calculate emotion distributions for direct, indirect, and early trials
        direct_emotion_distribution = direct_trials[emotion_columns].mean()
        indirect_emotion_distribution = indirect_trials[emotion_columns].mean()
        true_emotion_distribution = true_trials[emotion_columns].mean()
        false_emotion_distribution = false_trials[emotion_columns].mean()
        green_direct_early_distribution = green_direct_early_trials[emotion_columns].mean()
        red_direct_early_distribution = red_direct_early_trials[emotion_columns].mean()
        green_indirect_early_distribution = green_indirect_early_trials[emotion_columns].mean()
        red_indirect_early_distribution = red_indirect_early_trials[emotion_columns].mean()
        green_direct_late_distribution = green_direct_late_trials[emotion_columns].mean()
        red_direct_late_distribution = red_direct_late_trials[emotion_columns].mean()
        green_indirect_late_distribution = green_indirect_late_trials[emotion_columns].mean()
        red_indirect_late_distribution = red_indirect_late_trials[emotion_columns].mean()

        direct_row = pd.Series({'Trial': 'Direct Trials', **direct_emotion_distribution})
        indirect_row = pd.Series({'Trial': 'Indirect Trials', **indirect_emotion_distribution})
        true_row = pd.Series({'Trial': 'True Trials', **true_emotion_distribution})
        false_row = pd.Series({'Trial': 'False Trials', **false_emotion_distribution})
        green_direct_early_row = pd.Series({'Trial': 'Green Direct Early Trials', **green_direct_early_distribution})
        red_direct_early_row = pd.Series({'Trial': 'Red Direct Early Trials', **red_direct_early_distribution})
        green_indirect_early_row = pd.Series({'Trial': 'Green Indirect Early Trials', **green_indirect_early_distribution})
        red_indirect_early_row = pd.Series({'Trial': 'Red Indirect Early Trials', **red_indirect_early_distribution})
        green_direct_late_row = pd.Series({'Trial': 'Green Direct Late Trials', **green_direct_late_distribution})
        red_direct_late_row = pd.Series({'Trial': 'Red Direct Late Trials', **red_direct_late_distribution})
        green_indirect_late_row = pd.Series({'Trial': 'Green Indirect Late Trials', **green_indirect_late_distribution})
        red_indirect_late_row = pd.Series({'Trial': 'Red Indirect Late Trials', **red_indirect_late_distribution})

        green_direct_early_trials = early_trials[
            (early_trials['Cube Color'] == 'Green') & (early_trials['Correctness'] == 'direct')
        ]
        red_direct_early_trials = early_trials[
            (early_trials['Cube Color'] == 'Red') & (early_trials['Correctness'] == 'direct')
        ]
        green_indirect_early_trials = early_trials[
            (early_trials['Cube Color'] == 'Green') & (early_trials['Correctness'] == 'indirect')
        ]
        red_indirect_early_trials = early_trials[
            (early_trials['Cube Color'] == 'Red') & (early_trials['Correctness'] == 'indirect')
        ]
        green_direct_late_trials = late_trials[
            (late_trials['Cube Color'] == 'Green') & (late_trials['Correctness'] == 'direct')
        ]
        red_direct_late_trials = late_trials[
            (late_trials['Cube Color'] == 'Red') & (late_trials['Correctness'] == 'direct')
        ]
        green_indirect_late_trials = late_trials[
            (late_trials['Cube Color'] == 'Green') & (late_trials['Correctness'] == 'indirect')
        ]
        red_indirect_late_trials = late_trials[
            (late_trials['Cube Color'] == 'Red') & (late_trials['Correctness'] == 'indirect')
        ]

        # Perform True/False analysis for Green/Red Direct Early trials
        green_direct_early_true_trials = green_direct_early_trials[green_direct_early_trials['True/False'] == 'True']
        red_direct_early_true_trials = red_direct_early_trials[red_direct_early_trials['True/False'] == 'True']
        green_direct_early_false_trials = green_direct_early_trials[green_direct_early_trials['True/False'] == 'False']
        red_direct_early_false_trials = red_direct_early_trials[red_direct_early_trials['True/False'] == 'False']
        green_indirect_early_true_trials = green_indirect_early_trials[green_indirect_early_trials['True/False'] == 'True']
        red_indirect_early_true_trials = red_indirect_early_trials[red_indirect_early_trials['True/False'] == 'True']
        green_indirect_early_false_trials = green_indirect_early_trials[green_indirect_early_trials['True/False'] == 'False']
        red_indirect_early_false_trials = red_indirect_early_trials[red_indirect_early_trials['True/False'] == 'False']
        green_direct_late_true_trials = green_direct_late_trials[green_direct_late_trials['True/False'] == 'True']
        red_direct_late_true_trials = red_direct_late_trials[red_direct_late_trials['True/False'] == 'True']
        green_direct_late_false_trials = green_direct_late_trials[green_direct_late_trials['True/False'] == 'False']
        red_direct_late_false_trials = red_direct_late_trials[red_direct_late_trials['True/False'] == 'False']
        green_indirect_late_true_trials = green_indirect_late_trials[green_indirect_late_trials['True/False'] == 'True']
        red_indirect_late_true_trials = red_indirect_late_trials[red_indirect_late_trials['True/False'] == 'True']
        green_indirect_late_false_trials = green_indirect_late_trials[green_indirect_late_trials['True/False'] == 'False']
        red_indirect_late_false_trials = red_indirect_late_trials[red_indirect_late_trials['True/False'] == 'False']

        green_direct_early_true_distribution = green_direct_early_true_trials[emotion_columns].mean()
        red_direct_early_true_distribution = red_direct_early_true_trials[emotion_columns].mean()
        green_direct_early_false_distribution = green_direct_early_false_trials[emotion_columns].mean()
        red_direct_early_false_distribution = red_direct_early_false_trials[emotion_columns].mean()
        green_indirect_early_true_distribution = green_indirect_early_true_trials[emotion_columns].mean()
        red_indirect_early_true_distribution = red_indirect_early_true_trials[emotion_columns].mean()
        green_indirect_early_false_distribution = green_indirect_early_false_trials[emotion_columns].mean()
        red_indirect_early_false_distribution = red_indirect_early_false_trials[emotion_columns].mean()
        green_direct_late_true_distribution = green_direct_late_true_trials[emotion_columns].mean()
        red_direct_late_true_distribution = red_direct_late_true_trials[emotion_columns].mean()
        green_direct_late_false_distribution = green_direct_late_false_trials[emotion_columns].mean()
        red_direct_late_false_distribution = red_direct_late_false_trials[emotion_columns].mean()
        green_indirect_late_true_distribution = green_indirect_late_true_trials[emotion_columns].mean()
        red_indirect_late_true_distribution = red_indirect_late_true_trials[emotion_columns].mean()
        green_indirect_late_false_distribution = green_indirect_late_false_trials[emotion_columns].mean()
        red_indirect_late_false_distribution = red_indirect_late_false_trials[emotion_columns].mean()

        green_direct_early_true_row = pd.Series({'Trial': 'Green Direct Early True Trials', **green_direct_early_true_distribution})
        red_direct_early_true_row = pd.Series({'Trial': 'Red Direct Early True Trials', **red_direct_early_true_distribution})
        green_direct_early_false_row = pd.Series({'Trial': 'Green Direct Early False Trials', **green_direct_early_false_distribution})
        red_direct_early_false_row = pd.Series({'Trial': 'Red Direct Early False Trials', **red_direct_early_false_distribution})
        green_indirect_early_true_row = pd.Series({'Trial': 'Green Indirect Early True Trials', **green_indirect_early_true_distribution})
        red_indirect_early_true_row = pd.Series({'Trial': 'Red Indirect Early True Trials', **red_indirect_early_true_distribution})
        green_indirect_early_false_row = pd.Series({'Trial': 'Green Indirect Early False Trials', **green_indirect_early_false_distribution})
        red_indirect_early_false_row = pd.Series({'Trial': 'Red Indirect Early False Trials', **red_indirect_early_false_distribution})
        green_direct_late_true_row = pd.Series({'Trial': 'Green Direct Late True Trials', **green_direct_late_true_distribution})
        red_direct_late_true_row = pd.Series({'Trial': 'Red Direct Late True Trials', **red_direct_late_true_distribution})
        green_direct_late_false_row = pd.Series({'Trial': 'Green Direct Late False Trials', **green_direct_late_false_distribution})
        red_direct_late_false_row = pd.Series({'Trial': 'Red Direct Late False Trials', **red_direct_late_false_distribution})
        green_indirect_late_true_row = pd.Series({'Trial': 'Green Indirect Late True Trials', **green_indirect_late_true_distribution})
        red_indirect_late_true_row = pd.Series({'Trial': 'Red Indirect Late True Trials', **red_indirect_late_true_distribution})
        green_indirect_late_false_row = pd.Series({'Trial': 'Green Indirect Late False Trials', **green_indirect_late_false_distribution})
        red_indirect_late_false_row = pd.Series({'Trial': 'Red Indirect Late False Trials', **red_indirect_late_false_distribution})

        # descriptive label and NaN values
        separator_row = pd.Series({'Trial': '--- Separation for True/False Analysis ---', **{emotion: np.nan for emotion in emotion_columns}})
        trial_emotion_df = pd.concat([trial_emotion_df, pd.DataFrame([direct_row, indirect_row, true_row, false_row, green_direct_early_row, 
                                                                      red_direct_early_row, green_indirect_early_row, red_indirect_early_row,
                                                                      green_direct_late_row, red_direct_late_row, green_indirect_late_row, red_indirect_late_row,
                                                                      separator_row,
                                                                      green_direct_early_true_row, red_direct_early_true_row, green_direct_early_false_row, red_direct_early_false_row,
                                                                      green_indirect_early_true_row, red_indirect_early_true_row, green_indirect_early_false_row, red_indirect_early_false_row,
                                                                      green_direct_late_true_row, red_direct_late_true_row, green_direct_late_false_row, red_direct_late_false_row,
                                                                      green_indirect_late_true_row, red_indirect_late_true_row, green_indirect_late_false_row, red_indirect_late_false_row
                                                                      ])], ignore_index=True)

    return trial_emotion_df

excel_file = "block_analysis.xlsx"
# Process each block and save results to Excel
with pd.ExcelWriter(excel_file) as writer:
    for block_number, (block_start, block_end) in block_ranges.items():
        trial_emotion_df = process_block(block_number, block_start, block_end)

        # Write all analyses to the same sheet for each block
        trial_emotion_df.to_excel(writer, sheet_name=f"Block {block_number}", index=False)

print(f"Analysis saved to {excel_file}.")

# Apply formatting to make the first column bold for rows after row 17
wb = load_workbook(excel_file)
for block_number in block_ranges.keys():
    sheet = wb[f"Block {block_number}"]
    for row in sheet.iter_rows(min_row=18):  # Start formatting from row 18
        row[0].font = Font(bold=True)  # Apply bold formatting to the first column

wb.save(excel_file)
# print(f"Analysis saved to {excel_file} with first column bold after row 17.")